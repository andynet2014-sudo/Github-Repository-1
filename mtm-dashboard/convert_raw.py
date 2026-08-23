"""data/raw/ 의 증권사 원본 벌크 파일(xlsx)을 data/raw/csv/ 로 변환한다.

원본 xlsx는 그대로 보관(수기 대조·재확인용)하고, 여기서 만든 CSV는
읽기 쉬운 형태로 가공한 참고 자료다. collect.py가 관리하는 data/*.csv
(daily.csv, journal.csv 등)와는 별개이며 자동으로 병합되지 않는다 —
필요한 값은 사람이 확인 후 골라서 반영한다.

사용법: python3 convert_raw.py
"""

import csv
import re
from pathlib import Path

import openpyxl

RAW_DIR = Path(__file__).parent / "data" / "raw"
OUT_DIR = RAW_DIR / "csv"


def norm_num(v):
    """'811,800' 같은 콤마 포함 문자열 숫자를 숫자로 정규화."""
    if isinstance(v, str):
        s = v.replace(",", "").strip()
        if re.fullmatch(r"-?\d+(\.\d+)?", s):
            return float(s) if "." in s else int(s)
        return v
    return v


def convert_국장_거래내역():
    src = RAW_DIR / "거래내역_국장_2025-01-01_2026-08-21.xlsx"
    dst = OUT_DIR / "거래내역_국장_2025-01-01_2026-08-21.csv"
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb["거래내역"]
    header = [c.value for c in ws[4]]
    with dst.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(header)
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row[0] is None:
                continue
            w.writerow([norm_num(v) for v in row])
    print(f"wrote {dst}")


def convert_미장_거래내역():
    src = RAW_DIR / "거래내역_미장_2025-05-05_2026-08-21.xlsx"
    dst = OUT_DIR / "거래내역_미장_2025-05-05_2026-08-21.csv"
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb["거래내역"]
    header = [c.value for c in ws[1]]
    with dst.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(header)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            w.writerow([norm_num(v) for v in row])
    print(f"wrote {dst}")


def convert_미장_일별잔고():
    src = RAW_DIR / "일별잔고_미장_2026-01-01_2026-08-23.xlsx"
    dst = OUT_DIR / "일별잔고_미장_2026-01-01_2026-08-23.csv"
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb["계좌자산추이"]
    header = [c.value for c in ws[1]]
    with dst.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(header)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            date = row[0]
            date_str = date.strftime("%Y-%m-%d") if hasattr(date, "strftime") else date
            w.writerow([date_str] + [norm_num(v) for v in row[1:]])
    print(f"wrote {dst}")


if __name__ == "__main__":
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    convert_국장_거래내역()
    convert_미장_거래내역()
    convert_미장_일별잔고()
