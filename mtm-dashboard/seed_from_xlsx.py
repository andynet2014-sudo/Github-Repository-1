#!/usr/bin/env python3
"""v2.4.5 스프레드시트에서 collect.py 가 쓸 씨앗 데이터를 1회 추출한다.
원래 시트: 03_포지션 (초기 데이터 이관 전용, 1회성 부트스트랩 스크립트)

한 번만 돌리면 되고, 이후로는 data/*.csv 만 관리하면 된다.
사용법:  python3 seed_from_xlsx.py <v2.4.5.xlsx>
"""
import csv, os, sys, datetime
import openpyxl

DATA = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')


def quote_symbol(ticker):
    """시트의 종목코드를 시세 조회용 심볼로 변환."""
    if ticker.startswith('KRX:'):
        return ticker.split(':', 1)[1], 'KRW'
    return ticker, 'USD'


def main(path):
    os.makedirs(DATA, exist_ok=True)
    wb = openpyxl.load_workbook(path, data_only=True)

    # ---------- 보유 종목 ----------
    ws = wb['03_포지션']
    positions = []
    for r in range(17, 32):
        ticker = ws.cell(r, 1).value
        if not ticker:
            continue
        sym, ccy = quote_symbol(str(ticker))
        positions.append({
            'ticker': ticker, 'name': ws.cell(r, 2).value,
            'account': ws.cell(r, 3).value,
            'qty': ws.cell(r, 5).value,
            'avg_price_krw': round(float(ws.cell(r, 6).value), 4),
            'lev': int(ws.cell(r, 8).value or 1),
            'sector': ws.cell(r, 12).value,
            'uses_credit': 'Y' if str(ws.cell(r, 13).value).upper() == 'Y' else 'N',
            'stop_price_krw': round(float(ws.cell(r, 14).value), 4) if ws.cell(r, 14).value else '',
            'quote_symbol': sym, 'quote_ccy': ccy,
        })

    # ---------- 시세 캐시 (시트의 현재가를 초기값으로) ----------
    today = datetime.date.today().isoformat()
    prices = []
    for r in range(17, 32):
        ticker = ws.cell(r, 1).value
        if not ticker:
            continue
        prices.append({'ticker': ticker,
                       'price_krw': round(float(ws.cell(r, 7).value), 4),
                       'asof': today, 'source': 'seed:v2.4.5'})

    # ---------- 계좌 ----------
    accounts = [
        {'account': ws.cell(r, 1).value,
         'cash': ws.cell(r, 2).value or 0,
         'credit': ws.cell(r, 3).value or 0}
        for r in (6, 7, 8)
    ]
    common = {
        'company_loan': ws.cell(12, 2).value,      # 회사대출 잔액
        'margin_call_ratio': ws.cell(13, 2).value,  # 증권사 청산 담보비율
        'salary': ws.cell(14, 2).value,             # 연봉(세전)
    }

    # ---------- 룰북 임계값 ----------
    rb = wb['02_룰북']
    rules = []
    for r in range(6, 15):
        rid = rb.cell(r, 1).value
        if not rid:
            continue
        rules.append({'id': rid, 'metric': rb.cell(r, 3).value,
                      'threshold': rb.cell(r, 4).value,
                      'warn': rb.cell(r, 5).value,
                      'rule': rb.cell(r, 2).value})

    def dump(fn, rows, cols):
        with open(os.path.join(DATA, fn), 'w', newline='', encoding='utf-8') as f:
            w = csv.DictWriter(f, fieldnames=cols)
            w.writeheader()
            w.writerows(rows)
        print('  %-16s %d행' % (fn, len(rows)))

    print('씨앗 추출:')
    dump('positions.csv', positions, ['ticker', 'name', 'account', 'qty', 'avg_price_krw',
                                      'lev', 'sector', 'uses_credit', 'stop_price_krw',
                                      'quote_symbol', 'quote_ccy'])
    dump('prices.csv', prices, ['ticker', 'price_krw', 'asof', 'source'])
    dump('accounts.csv', accounts, ['account', 'cash', 'credit'])
    dump('rules.csv', rules, ['id', 'metric', 'threshold', 'warn', 'rule'])
    with open(os.path.join(DATA, 'common.csv'), 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['key', 'value'])
        for k, v in common.items():
            w.writerow([k, v])
    print('  %-16s %d행' % ('common.csv', len(common)))


if __name__ == '__main__':
    main(sys.argv[1] if len(sys.argv) > 1 else 'v2.4.5.xlsx')
