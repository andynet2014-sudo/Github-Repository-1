#!/usr/bin/env python3
"""리스크 대시보드 수집기 — 매일 입력을 0으로 만든다.
원래 시트: 00_대시보드 + 02_룰북 + 03_포지션 (핵심 계산 엔진, 매일 실행)

사람이 만지는 파일은 두 개뿐이고, 둘 다 '거래한 날'에만 고치면 된다.
  data/positions.csv   종목·수량·평단·손절가·섹터
  data/accounts.csv    예수금·신용잔고

나머지(prices/daily)는 이 스크립트가 쓴다.

지표 정의는 260731_리스크관리대시보드_v2.4.5 의 '③ 포지션 ④ 리스크 엔진'과
동일하다. 계산 위치만 스프레드시트에서 파이썬으로 옮긴 것이다.

  python3 collect.py                 오늘자 수집·기록
  python3 collect.py --dry           계산만 하고 저장하지 않음
  python3 collect.py --no-fetch      시세 조회 없이 캐시로 재계산
  python3 collect.py --date 2026-08-12
  python3 collect.py --backfill v2.4.5.xlsx    01_일별로그 과거 이력 복원

이 파일은 오케스트레이터(main)와 핵심 계산(compute)만 담당한다. 시세 조회는
pricing.py, 매크로 지표는 macro.py, 룰 판정은 rules.py, 콘솔 출력은 report.py에
있다 — 여러 세션이 동시에 서로 다른 기능을 고칠 때 이 파일 하나에 다 몰려있어서
git 충돌이 잦았던 문제를 줄이기 위해 2026-09 분리했다(DECISIONS.md 참고).
기존에 `import collect; collect.compute(...)`처럼 쓰던 다운스트림 스크립트가
안 깨지도록, 분리된 모듈의 이름들을 아래에서 그대로 재노출(re-export)한다.
"""
import argparse
import datetime

from base import BASE, DATA, num, read_csv, write_csv
from pricing import (NONMARKET, PRICE_HISTORY_COLS, fetch_krx, fetch_yahoo,
                     fetch_yahoo_history, instrument_list, fetch_price_table,
                     upsert_price_history, backfill_prices, update_positions_current_price)
from macro import (MACRO_TICKERS, MACRO_OHLC_SUFFIXES, MACRO_COLS,
                   load_macro, upsert_macro, backfill_macro)
from rules import RULE_MAP, FALLBACK_THRESHOLD, judge
from report import PCT, MULT, fmt, report
from investor_flow import (INVESTOR_MARKET_KEYS, INVESTOR_MARKET_COLS, INVESTOR_STOCK_COLS,
                           STOCK_FLOW_TICKERS, fetch_investor_flow_market,
                           upsert_investor_flow_market, fetch_investor_flow_stock_recent,
                           upsert_investor_flow_stock, collect_investor_flow)


def load_cashflow_totals(common):
    """cashflow.csv 가 있으면 원금·이자·강의비 누계를 여기서 직접 합산해 common 을
    덮어쓴다. common.csv 를 손으로 따로 갱신할 필요가 없어지고, cashflow.csv 에
    행만 추가하면 다음 실행부터 자동으로 반영된다 — 두 파일이 따로 놀며 어긋나는 걸
    막기 위한 단일 소스."""
    rows = read_csv('cashflow.csv')
    if not rows:
        return common
    common = dict(common)
    NON_PRINCIPAL = ('신용이자', '강의구독비', '수수료', '매도세금', '배당금')
    interest = sum(num(r['amount']) for r in rows if r['type'] == '신용이자')
    course_fees = sum(num(r['amount']) for r in rows if r['type'] == '강의구독비')
    trade_fees = sum(num(r['amount']) for r in rows if r['type'] == '수수료')
    sec_tax = sum(num(r['amount']) for r in rows if r['type'] == '매도세금')
    dividend = sum(num(r['amount']) for r in rows if r['type'] == '배당금')
    principal = {'국장': 0.0, '미장': 0.0, 'ISA': 0.0}
    for r in rows:
        if r['type'] in NON_PRINCIPAL:
            continue
        amt = num(r['amount'])
        if r['from_account'] in principal:
            principal[r['from_account']] -= amt
        if r['to_account'] in principal:
            principal[r['to_account']] += amt
    common['cum_interest'] = interest
    common['cum_course_fees'] = course_fees
    common['cum_trade_fees'] = trade_fees
    common['cum_sec_tax'] = sec_tax
    common['cum_dividend'] = dividend
    for acct in ('국장', '미장', 'ISA'):
        common['principal_' + acct] = principal[acct]
    common['principal_total'] = sum(principal.values())
    return common


# ─────────────────────────── 리스크 엔진 ───────────────────────────

ACCOUNTS = ('국장', '미장', 'ISA')


def compute(positions, prices, common):
    """v2.4.5 '④ 리스크 엔진' 과 같은 정의.

    positions.csv 의 sector == '현금'/'신용' 행은 주식이 아니다 — 명목·실질
    Exposure·섹터·손절 계산에서 전부 제외하고, 계좌별로 cash/credit 에 합산한다.
    (03_포지션 탭의 '① 계좌별 입력'이 하던 역할을 이제 positions.csv 행이 대신한다.
    accounts.csv 는 삭제됐다 — 신용도 거래한 날 만지는 positions.csv 안에 있다.)
    """
    all_rows = []
    for p in positions:
        price = prices[p['ticker']][0]
        qty, lev = num(p['qty']), int(num(p['lev'], 1))
        nominal = qty * price               # I열 평가금액(명목)
        all_rows.append(dict(p, price=price, qty=qty, lev=lev,
                             nominal=nominal, real=nominal * lev))   # J열 (실질)

    rows = [r for r in all_rows if r['sector'] not in ('현금', '신용')]
    cash = {a: 0.0 for a in ACCOUNTS}
    credit = {a: 0.0 for a in ACCOUNTS}
    for r in all_rows:
        if r['sector'] == '현금':
            cash[r['account']] = cash.get(r['account'], 0.0) + r['nominal']
        elif r['sector'] == '신용':
            credit[r['account']] = credit.get(r['account'], 0.0) + r['nominal']

    nominal_sum = sum(r['nominal'] for r in rows)          # B36 주식평가(명목)
    expo_real = sum(r['real'] for r in rows)               # B37 주식 Exposure(실질)
    cash_sum = sum(cash.values())                          # B38 예수금 합계
    total_val = nominal_sum + cash_sum                     # B39 총평가(명목)
    credit_sum = sum(credit.values())                      # B40 증권사 신용
    equity = total_val - credit_sum                        # B41 계좌 순자산
    loan = num(common['company_loan'])                     # B42 회사대출
    own = equity - loan                                    # B43 순수 자기자본
    implied = sum(r['nominal'] * (r['lev'] - 1) for r in rows)   # B44 내재 레버리지
    debt = credit_sum + implied                            # B45 총차입

    kr = [r for r in rows if r['account'] == '국장']
    kr_total = sum(r['nominal'] for r in kr) + cash.get('국장', 0)   # B46
    kr_expo = sum(r['real'] for r in kr)                             # B47
    kr_credit = credit.get('국장', 0)
    mcr = num(common['margin_call_ratio'])

    # 섹터 집계 — 실질 기준, 분모는 실질 합계
    sector = {}
    for r in rows:
        sector[r['sector']] = sector.get(r['sector'], 0) + r['real']
    top_sector, top_sector_val = max(sector.items(), key=lambda x: x[1]) if sector else ('', 0)

    # 손절: 현재가가 손절가 위인 종목만 손실이 잡힌다 (아래면 이미 하회 → 0)
    # stop_price_krw 가 avg_price_krw 와 사실상 같으면(0.1% 이내) '평단을 손절가 칸에
    # 복사만 해둔 것' — 실손절가를 정한 게 아니다. 그런 값을 진짜 손절가로 세면
    # R04/R05 가 "손절 관리 잘 되고 있음"으로 잘못 읽힌다 (README 참고: 8/9종목이 이 상태).
    # 그래서 stop_missing(공란)과 별도로 stop_at_cost(본전=손절가, 사실상 미설정)를 센다.
    def has_real_stop(r):
        sp = r['stop_price_krw']
        if sp in ('', None):
            return False
        avg = num(r['avg_price_krw'])
        return not (avg and abs(num(sp) - avg) / avg < 0.001)

    stop_missing = sum(1 for r in rows if r['stop_price_krw'] in ('', None))
    stop_at_cost = sum(1 for r in rows
                       if r['stop_price_krw'] not in ('', None) and not has_real_stop(r))
    stop_below = sum(1 for r in rows
                     if has_real_stop(r) and r['price'] < num(r['stop_price_krw']))
    stop_loss = sum(max(0.0, (r['price'] - num(r['stop_price_krw'])) * r['qty'])
                    for r in rows if has_real_stop(r))

    lev_etf_nominal = sum(r['nominal'] for r in rows if r['lev'] > 1)
    salary = num(common['salary'])

    def div(a, b):
        return a / b if b else float('nan')

    # 09_CF_원금 Ⅴ — 비용 차감 후 실질 손익 (이자는 참고만, 차감하지 않는다)
    principal_total = num(common.get('principal_total', 0))
    surface_pnl = equity - principal_total
    course_fees = num(common.get('cum_course_fees', 0))
    tax_est = num(common.get('cum_tax_est', 0))
    net_pnl = surface_pnl - course_fees - tax_est

    m = {
        'nominal_sum': nominal_sum, 'expo_real': expo_real, 'cash': cash_sum,
        'total_val': total_val, 'credit': credit_sum, 'equity': equity,
        'loan': loan, 'own_equity': own, 'implied_lev': implied, 'debt': debt,
        'kr_total': kr_total, 'kr_expo': kr_expo,
        'margin_ratio': div(kr_total, kr_credit),                       # B51
        'liq_room': div(kr_total - kr_credit * mcr, kr_expo),           # B52
        'lev_account': div(expo_real, equity),                          # B53
        'lev_real': div(expo_real, own),                                # B54
        'debt_ratio': div(debt, expo_real),                             # B55
        'cash_ratio': div(cash_sum, expo_real + cash_sum),              # B56
        'n_positions': len(rows),                                       # B57
        'max_pos': div(max((r['real'] for r in rows), default=0), nominal_sum),   # B58
        'max_sector': div(top_sector_val, expo_real),                   # B59
        'max_sector_name': top_sector,
        'stop_missing': stop_missing, 'stop_below': stop_below,         # B60/B61
        'stop_at_cost': stop_at_cost,          # 본전가=손절가로 등록된, 사실상 미설정 종목 수
        'n_credit': sum(1 for r in rows if r['uses_credit'] == 'Y'),    # B62
        'stop_loss': stop_loss,                                         # B63
        'stop_loss_ratio': div(stop_loss, own),                         # B64
        'debt_to_salary': div(debt, salary),                            # B65
        'lev_etf_ratio': div(lev_etf_nominal, equity),                  # R07
        'lev_etf_nominal': lev_etf_nominal,
        'principal_total': principal_total, 'surface_pnl': surface_pnl,
        'cum_interest': num(common.get('cum_interest', 0)),
        'cum_trade_fees': num(common.get('cum_trade_fees', 0)),
        'cum_sec_tax': num(common.get('cum_sec_tax', 0)),
        'cum_course_fees': course_fees, 'cum_tax_est': tax_est,
        'net_pnl': net_pnl, 'net_return': div(net_pnl, principal_total),
    }
    for acct in ('국장', '미장', 'ISA'):
        sub = [r for r in rows if r['account'] == acct]
        m[acct + '_total'] = sum(r['nominal'] for r in sub) + cash.get(acct, 0)
        m[acct + '_equity'] = m[acct + '_total'] - credit.get(acct, 0)
    return m, rows


# ─────────────────────────── daily.csv ───────────────────────────

DAILY_COLS = ['date', 'equity', 'own_equity', 'expo_real', 'nominal_sum', 'cash',
              'credit', 'implied_lev', 'debt', 'lev_real', 'lev_account',
              'margin_ratio', 'liq_room', 'cash_ratio', 'debt_ratio',
              'max_sector', 'max_sector_name', 'max_pos', 'stop_below',
              'stop_missing', 'stop_at_cost', 'stop_loss', 'stop_loss_ratio', 'debt_to_salary',
              'surface_pnl', 'net_pnl', 'net_return',
              'violations', '국장_total', '국장_equity', '미장_equity', 'ISA_equity',
              'emotion', 'source']


def upsert_daily(date, m, signals, source, emotion=None):
    existing = {r['date']: r for r in read_csv('daily.csv')}
    row = {'date': date, 'source': source,
           'emotion': emotion if emotion is not None else existing.get(date, {}).get('emotion', ''),
           'violations': ','.join(s[0] for s in signals if s[4].startswith('🔴'))}
    for k in DAILY_COLS:
        if k in row:
            continue
        v = m.get(k)
        if isinstance(v, float):
            v = round(v, 6)
        row[k] = '' if v is None else v
    rows = [r for r in read_csv('daily.csv') if r['date'] != date]
    rows.append(row)
    rows.sort(key=lambda r: r['date'])
    write_csv('daily.csv', rows, DAILY_COLS)
    return row


# ─────────────────────────── positions_history.csv ───────────────────────────
# daily.csv 는 계좌 합계만 남긴다 — 종목별로 그날 얼마였는지는 여기 쌓인다.
# positions.csv 는 최신 상태로 매번 덮어써지므로 그 자체로는 이력이 안 남는다.

POSITION_HISTORY_COLS = ['date', 'ticker', 'name', 'account', 'sector', 'qty', 'lev',
                          'avg_price_krw', 'price', 'nominal', 'real', 'pnl', 'pnl_pct']


def upsert_position_history(date, rows):
    """positions.csv 의 그날 스냅샷(종목별 가격·평가·손익)을 date+ticker+account 키로 쌓는다."""
    keep = [r for r in read_csv('positions_history.csv')
            if not (r['date'] == date)]
    for r in rows:
        avg = num(r['avg_price_krw'])
        market = r['sector'] not in ('현금', '신용')
        pnl = (r['price'] - avg) * r['qty'] if market else 0.0
        pnl_pct = (r['price'] / avg - 1) if avg and market else 0.0
        keep.append({
            'date': date, 'ticker': r['ticker'], 'name': r['name'], 'account': r['account'],
            'sector': r['sector'], 'qty': r['qty'], 'lev': r['lev'],
            'avg_price_krw': avg, 'price': r['price'], 'nominal': r['nominal'], 'real': r['real'],
            'pnl': round(pnl, 2), 'pnl_pct': round(pnl_pct, 6),
        })
    keep.sort(key=lambda r: (r['date'], r['account'], r['ticker']))
    write_csv('positions_history.csv', keep, POSITION_HISTORY_COLS)


# ─────────────────────────── backfill ───────────────────────────

def backfill(xlsx):
    """01_일별로그의 과거 입력값을 daily.csv 로 복원.

    2026-07-30 이전에는 미장·ISA 기록이 아예 없다. 그 구간에서 국장만 더한 값을
    '전체 순자산'으로 적으면 추이가 조용히 틀리므로, 국장 값만 채우고 전체
    지표는 비워 둔 뒤 source 로 구분한다.

    S~V(일별 빠른입력, 19~22열)가 없는 날도 B열(순자산)·D열(누적손익)이 채워져
    있으면 그걸로 채운다 — 매일은 아니어도 주기적으로 기록해 둔 과거 시점들이다.
    누적손익(D열)은 시트 전체에서 '순자산 − 340,000,000(고정 원금)'으로 일관돼 있어
    collect.py 의 surface_pnl 과 같은 정의다.
    """
    import openpyxl
    ws = openpyxl.load_workbook(xlsx, data_only=True)['01_일별로그']
    existing = {r['date']: r for r in read_csv('daily.csv')}
    added = 0
    kinds = {}
    for r in range(5, 1000):
        d = ws.cell(r, 1).value
        if not isinstance(d, datetime.datetime):
            continue
        date = d.strftime('%Y-%m-%d')
        if date in existing:
            continue
        S, T = ws.cell(r, 19).value, ws.cell(r, 20).value      # 국장 총평가 / 순자산
        U, V = ws.cell(r, 21).value, ws.cell(r, 22).value      # 미장 / ISA 순자산
        asset, cum_pnl = ws.cell(r, 2).value, ws.cell(r, 4).value   # 순자산(B) / 누적손익(D)
        if not any((S, T, U, V, asset)):
            continue
        row = {c: '' for c in DAILY_COLS}
        row['date'] = date
        row['국장_total'] = S or ''
        row['국장_equity'] = T or ''
        row['미장_equity'] = U or ''
        row['ISA_equity'] = V or ''
        if T and U and V:
            row['equity'] = num(T) + num(U) + num(V)
            kind = 'backfill(전체)'
        elif T and not (U or V):
            kind = 'backfill(국장만)'          # 미장·ISA 기록이 아예 없던 구간
        elif asset:
            row['equity'] = num(asset)
            kind = 'backfill(순자산)'          # S~V 없이 주기적 순자산 스냅샷만 있는 구간
        else:
            kind = 'backfill(국장순자산 누락)'  # 미장·ISA 는 있는데 국장 순자산이 빈 행
        if cum_pnl is not None:
            row['surface_pnl'] = num(cum_pnl)
        row['source'] = kind
        kinds[kind] = kinds.get(kind, 0) + 1
        existing[date] = row
        added += 1
    rows = sorted(existing.values(), key=lambda x: x['date'])
    write_csv('daily.csv', rows, DAILY_COLS)
    print('복원 %d행 (daily.csv 총 %d행)' % (added, len(rows)))
    for k, v in sorted(kinds.items()):
        print('  %-26s %d행' % (k, v))
    if kinds.get('backfill(전체)', 0) < added:
        print('\n  ※ equity(전체 순자산)는 세 계좌가 모두 있는 행에만 채웠습니다.')
        print('    나머지 행에 국장 값만 넣어 "전체 순자산"으로 그리면 추이가 조용히 틀립니다.')


# ─────────────────────────── main ───────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--dry', action='store_true', help='계산만, 저장 안 함')
    ap.add_argument('--no-fetch', action='store_true', help='시세 조회 없이 캐시 사용')
    ap.add_argument('--date', help='기록 날짜 (기본: 오늘)')
    ap.add_argument('--emotion', help='오늘의 감정 태그 (예: 탐욕, 공포, 평온)')
    ap.add_argument('--backfill', metavar='XLSX', help='01_일별로그 과거 이력 복원')
    ap.add_argument('--backfill-prices', metavar='START_DATE',
                    help='이 날짜부터 오늘까지 종목 종가를 야후에서 소급 조회 (예: 2026-08-01)')
    ap.add_argument('--backfill-prices-end', metavar='END_DATE', help='소급 조회 종료일 (기본: 오늘)')
    ap.add_argument('--backfill-macro', metavar='START_DATE',
                    help='이 날짜부터 오늘까지 매크로 지표를 야후에서 소급 조회 (예: 2026-01-01)')
    ap.add_argument('--backfill-macro-end', metavar='END_DATE', help='소급 조회 종료일 (기본: 오늘)')
    ap.add_argument('--investor-flow', metavar='DATE', nargs='?', const='today',
                    help='코스피 전체 + 삼전/SK하닉 투자자 순매수만 조회(단독 실행, 예: 2026-09-03)')
    a = ap.parse_args()

    if a.backfill:
        return backfill(a.backfill)
    if a.backfill_prices:
        return backfill_prices(a.backfill_prices, a.backfill_prices_end)
    if a.backfill_macro:
        return backfill_macro(a.backfill_macro, a.backfill_macro_end)
    if a.investor_flow:
        d = datetime.date.today().isoformat() if a.investor_flow == 'today' else a.investor_flow
        ok, failed = collect_investor_flow(d)
        print('성공:', ok)
        print('실패:', failed)
        return

    date = a.date or datetime.date.today().isoformat()
    positions = read_csv('positions.csv')
    if not positions:
        raise SystemExit('data/positions.csv 가 없습니다. seed_from_xlsx.py 를 먼저 돌리세요.')
    common = {r['key']: r['value'] for r in read_csv('common.csv')}
    common = load_cashflow_totals(common)
    rules = read_csv('rules.csv')

    instruments = instrument_list(positions)
    prices, failed, price_rows = fetch_price_table(instruments, not a.no_fetch, date)
    for p in positions:                     # 현금/신용은 시장 시세가 없다 — 사람이 적은 값 그대로
        if p['sector'] in NONMARKET:
            prices[p['ticker']] = (num(p['current_price_krw'] or p['avg_price_krw']),
                                   'manual(%s)' % p['sector'])

    m, stock_rows = compute(positions, prices, common)
    signals = judge(m, rules)
    macro, macro_failed = load_macro(not a.no_fetch, date)
    report(date, m, signals, failed + macro_failed, macro)

    if a.dry:
        print('  (--dry: 저장하지 않았습니다)\n')
        return
    # prices.csv 는 export_to_xlsx.py/rebuild_dashboard_reference.py 가 읽는
    # "티커→최신 원화가" 스냅샷 포맷을 그대로 유지한다 (price_history.csv 와는 별개).
    write_csv('prices.csv',
              [{'ticker': t, 'price_krw': v[0], 'asof': date, 'source': v[1]}
               for t, v in prices.items()],
              ['ticker', 'price_krw', 'asof', 'source'])
    upsert_price_history(date, price_rows)
    update_positions_current_price(prices)
    upsert_daily(date, m, signals, 'live' if not a.no_fetch else 'cache', emotion=a.emotion)
    if macro:
        upsert_macro(date, macro)
    if not a.no_fetch:
        flow_ok, flow_failed = collect_investor_flow(date)
        if flow_ok:
            print('  투자자 순매수 수집: %s' % ', '.join(flow_ok))
        for name, err in flow_failed:
            print('  투자자 순매수 실패: %s (%s)' % (name, err))

    nonmarket_rows = []
    for p in positions:
        if p['sector'] not in NONMARKET:
            continue
        price = prices[p['ticker']][0]
        qty = num(p['qty'])
        nonmarket_rows.append(dict(p, price=price, qty=qty, lev=1,
                                    nominal=qty * price, real=qty * price))
    upsert_position_history(date, stock_rows + nonmarket_rows)

    print('  기록 완료 → data/daily.csv, data/macro.csv, data/price_history.csv, data/prices.csv, data/positions.csv(현재가), data/positions_history.csv\n')


if __name__ == '__main__':
    main()
