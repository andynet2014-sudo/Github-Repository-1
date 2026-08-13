#!/usr/bin/env python3
"""오늘자(또는 지정 날짜) 데이터를 v2.4.5 대시보드 스타일 엑셀 파일로 출력한다.
원래 시트: 00_대시보드 + 01_일별로그 (정적 엑셀 스냅샷 출력 전용)

collect.py 와 같은 compute()/judge() 를 그대로 써서, 화면에 뜨는 숫자와
엑셀 파일의 숫자가 항상 일치한다.

  python3 export_to_xlsx.py                 오늘자 스냅샷
  python3 export_to_xlsx.py --date 2026-08-11
"""
import argparse, datetime, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import collect

RED = PatternFill('solid', fgColor='FFE5E5')
GREEN = PatternFill('solid', fgColor='E5F7E5')
GRAY = PatternFill('solid', fgColor='EFEFEF')
HEADER = PatternFill('solid', fgColor='2B2F38')
HEADER_FONT = Font(color='FFFFFF', bold=True)
BOLD = Font(bold=True)
BIG = Font(bold=True, size=14)
THIN = Side(style='thin', color='CCCCCC')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def tile(ws, row, col, label, value, fmt=None):
    c1 = ws.cell(row, col, label)
    c1.font = Font(size=9, color='666666')
    c2 = ws.cell(row + 1, col, value)
    c2.font = BIG
    if fmt:
        c2.number_format = fmt
    for r in (row, row + 1):
        ws.cell(r, col).border = BOX


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--date')
    a = ap.parse_args()
    date = a.date or datetime.date.today().isoformat()

    positions = collect.read_csv('positions.csv')
    common = {r['key']: r['value'] for r in collect.read_csv('common.csv')}
    rules = collect.read_csv('rules.csv')
    prices, _, _ = collect.load_prices(positions, False, date)
    m, _ = collect.compute(positions, prices, common)
    signals = collect.judge(m, rules)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '대시보드'
    ws.sheet_view.showGridLines = False
    for col, w in zip('ABCDEFGH', [16, 16, 16, 16, 12, 12, 12, 12]):
        ws.column_dimensions[col].width = w

    ws.merge_cells('A1:H1')
    ws['A1'] = '리스크 대시보드 — %s' % date
    ws['A1'].font = Font(bold=True, size=16)

    labels = ['계좌 순자산', '순수 자기자본', '실질 Exposure', '총차입',
              '실질 레버리지', '청산까지 여력', '현금 비중', '최대 섹터']
    values = [m['equity'], m['own_equity'], m['expo_real'], m['debt'],
              m['lev_real'], m['liq_room'], m['cash_ratio'], m['max_sector']]
    fmts = ['#,##0', '#,##0', '#,##0', '#,##0', '0.00"배"', '0.0%', '0.0%', '0.0%']
    for i, (lb, v, f) in enumerate(zip(labels, values, fmts)):
        tile(ws, 3, 1 + i, lb, v, f)

    r = 6
    ws.cell(r, 1, '계좌별 현황').font = BOLD
    r += 1
    for j, h in enumerate(['계좌', '총평가', '순자산']):
        c = ws.cell(r, 1 + j, h)
        c.fill = HEADER
        c.font = HEADER_FONT
    r += 1
    for acct in ('국장', '미장', 'ISA'):
        ws.cell(r, 1, acct)
        ws.cell(r, 2, m[acct + '_total']).number_format = '#,##0'
        ws.cell(r, 3, m[acct + '_equity']).number_format = '#,##0'
        for c in range(1, 4):
            ws.cell(r, c).border = BOX
        r += 1

    r += 1
    ws.cell(r, 1, '리스크 신호등').font = BOLD
    r += 1
    for j, h in enumerate(['ID', '항목', '현재값', '기준값', '신호']):
        c = ws.cell(r, 1 + j, h)
        c.fill = HEADER
        c.font = HEADER_FONT
    r += 1
    for rid, label, cur, thr, sig, fallback in signals:
        key = collect.RULE_MAP[rid][0]
        ws.cell(r, 1, rid)
        ws.cell(r, 2, label)
        ws.cell(r, 3, collect.fmt(key, cur))
        ws.cell(r, 4, collect.fmt(key, thr) if thr is not None else '—')
        cell = ws.cell(r, 5, sig)
        cell.fill = RED if '🔴' in sig else GREEN
        for c in range(1, 6):
            ws.cell(r, c).border = BOX
        r += 1

    # 추이 시트 — daily.csv 그대로
    ws2 = wb.create_sheet('추이(daily.csv)')
    daily_rows = collect.read_csv('daily.csv')
    if daily_rows:
        headers = list(daily_rows[0].keys())
        for j, h in enumerate(headers):
            c = ws2.cell(1, 1 + j, h)
            c.fill = HEADER
            c.font = HEADER_FONT
        for i, row in enumerate(daily_rows):
            for j, h in enumerate(headers):
                v = row[h]
                try:
                    v = float(v) if v not in ('', None) else None
                except ValueError:
                    pass
                ws2.cell(2 + i, 1 + j, v)
        ws2.freeze_panes = 'B2'

    out = 'snapshot_%s.xlsx' % date
    wb.save(out)
    print('저장 완료 →', out)


if __name__ == '__main__':
    main()
